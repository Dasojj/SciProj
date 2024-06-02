import telebot
from telebot import types
import os
import openpyxl
import random
import re

bot = telebot.TeleBot('7179760881:AAEXuaVcaM1NGrDyOL4ZzfmWARNmPIEn8YQ')

current_selection = {}
user_state = {}
admin_codes = []

manuals_folder = "manuals"
variants_folder = "variants"
submissions_folder = "submissions"
checked_folder = "checked"
state_file = "user_state.conf"
admin_codes_file = "admin_codes.txt"

def load_admin_codes():
    if os.path.exists(admin_codes_file):
        with open(admin_codes_file, 'r') as f:
            for line in f:
                admin_codes.append(line.strip())

def save_user_state():
    with open(state_file, 'w') as f:
        for chat_id, data in user_state.items():
            f.write(f"{chat_id},{data['group']},{data['name']},{data.get('admin', False)}\n")

def load_user_state():
    if os.path.exists(state_file):
        with open(state_file, 'r') as f:
            for line in f:
                parts = line.strip().split(',')
                chat_id, group, name = parts[0], parts[1], parts[2]
                admin = parts[3].lower() == 'true' if len(parts) > 3 else False
                user_state[int(chat_id)] = {"group": group, "name": name, "admin": admin}

def send_pdf(chat_id, pdf_folder, pdf_name):
    pdf_path = os.path.join(pdf_folder, pdf_name)
    with open(pdf_path, 'rb') as f:
        bot.send_document(chat_id, f)

def find_variant(group, name):
    excel_file = os.path.join("groups", group + ".xlsx")
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == name:
                return row[1]
        return None
    else:
        return None

def update_variant(group, name, variant):
    excel_file = os.path.join("groups", group + ".xlsx")
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == name:
                row[1].value = variant
                wb.save(excel_file)
                return True
        return False
    else:
        return False
    
def group_exists(group):
    excel_file = os.path.join("groups", group + ".xlsx")
    return os.path.exists(excel_file)

def name_exists_in_group(group, name):
    excel_file = os.path.join("groups", group + ".xlsx")
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == name:
                return True
    return False

def start_message(message, edit=False):
    chat_id = message.chat.id
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    if chat_id in user_state and user_state[chat_id].get("admin"):
        btn_tables = types.InlineKeyboardButton(text="Таблицы", callback_data="tables")
        btn_check = types.InlineKeyboardButton(text="Проверка", callback_data="check")
        keyboard.add(btn_tables, btn_check)
        msg_text = "Выберите действие:"
        bot.send_message(chat_id, msg_text, reply_markup=keyboard)
    else:
        btn_variants = types.InlineKeyboardButton(text="Вариант", callback_data="variants")
        btn_materials = types.InlineKeyboardButton(text="Материалы", callback_data="materials")
        btn_submission = types.InlineKeyboardButton(text="Сдача", callback_data="submission")
        keyboard.add(btn_variants, btn_materials, btn_submission)
        msg_text = "Выберите один из вариантов:"
        if edit:
            bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=msg_text, reply_markup=keyboard)
        else:
            bot.send_message(chat_id, msg_text, reply_markup=keyboard)

@bot.message_handler(commands=['start'])
def handle_start(message):
    start_message(message)

@bot.message_handler(commands=['admin'])
def handle_admin(message):
    chat_id = message.chat.id
    msg = bot.send_message(chat_id, "Введите админ-код:")
    bot.register_next_step_handler(msg, process_admin_code)

@bot.message_handler(commands=['exit_admin'])
def handle_exit_admin(message):
    chat_id = message.chat.id
    if chat_id in user_state and user_state[chat_id].get("admin"):
        user_state[chat_id]["admin"] = False
        save_user_state()
        bot.send_message(chat_id, "Вы вышли из админ-режима.")
        start_message(message)
    else:
        bot.send_message(chat_id, "Вы не находитесь в админ-режиме.")

def process_admin_code(message):
    chat_id = message.chat.id
    code = message.text.strip()
    if code in admin_codes:
        user_state[chat_id] = user_state.get(chat_id, {})
        if (not user_state[chat_id]["group"]) or (not user_state[chat_id]["name"]):
            user_state[chat_id]["group"] = " "
            user_state[chat_id]["name"] = " "
        user_state[chat_id]["admin"] = True
        save_user_state()
        bot.send_message(chat_id, "Админ-режим активирован.")
        start_message(message)
    else:
        bot.send_message(chat_id, "Неверный админ-код.")
        start_message(message)

@bot.callback_query_handler(func=lambda call: True)
def handle_buttons(call):
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    if call.data == "variants":
        if chat_id in user_state and "group" in user_state[chat_id] and "name" in user_state[chat_id]:
            group = user_state[chat_id]["group"]
            name = user_state[chat_id]["name"]
            variant = find_variant(group, name)
            if variant is not None:
                pdf_name = f"Вариант_{variant}.pdf"
                send_pdf(chat_id, variants_folder, pdf_name)
                start_message(call.message)
                bot.delete_message(chat_id, message_id)
                return
        bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="Введите вашу группу (например, М8О-305Б-21):")
        bot.register_next_step_handler(call.message, process_group_step)
    elif call.data == "materials":
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        btn1 = types.InlineKeyboardButton(text="С разделяющимися переменными", callback_data="Пособие_1")
        btn2 = types.InlineKeyboardButton(text="Однородные уравнения", callback_data="Пособие_2")
        btn3 = types.InlineKeyboardButton(text="Линейные уравнения первого порядка", callback_data="Пособие_3")
        btn4 = types.InlineKeyboardButton(text="Уравнения в полных дифференциалах", callback_data="Пособие_4")
        btn5 = types.InlineKeyboardButton(text="Допускающие понижение порядка", callback_data="Пособие_5")
        btn6 = types.InlineKeyboardButton(text="Однородные относительно искомой функции", callback_data="Пособие_6")
        btn7 = types.InlineKeyboardButton(text="Со специальной правой частью", callback_data="Пособие_7")
        btn8 = types.InlineKeyboardButton(text="Вариация произвольных постоянных", callback_data="Пособие_8")
        keyboard.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8)
        bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="Выберите тему:", reply_markup=keyboard)
    elif call.data == "submission":
        bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="Введите номер задачи (от 1 до 15):")
        bot.register_next_step_handler(call.message, process_task_number)
    elif call.data == "tables":
        bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="Введите название группы для скачивания таблицы:")
        bot.register_next_step_handler(call.message, process_group_download)
    elif call.data == "check":
        process_check_groups(call.message)
    elif call.data.startswith("group_"):
        group_name = call.data.split("_")[1]
        process_check_students(call.message, group_name)
    elif call.data.startswith("student_"):
        student_name = call.data.split("_")[1]
        process_check_tasks(call.message, student_name)
    elif call.data.startswith("task_"):
        task_number = call.data.split("_")[1]
        process_send_task(call.message, task_number)
    elif call.data.startswith("pass_"):
        task_number = call.data.split("_")[1]
        group_name = current_selection[chat_id]['group']
        student_name = current_selection[chat_id]['student']
        update_task_status(group_name, student_name, task_number, 1)
        bot.send_message(chat_id, "Задача зачтена.")
        notify_student(group_name, student_name, task_number, "Зачёт")
        move_submission_to_checked(group_name, student_name, task_number)
        start_message(call.message)
    elif call.data.startswith("fail_"):
        task_number = call.data.split("_")[1]
        group_name = current_selection[chat_id]['group']
        student_name = current_selection[chat_id]['student']
        update_task_status(group_name, student_name, task_number, 0)
        bot.send_message(chat_id, "Задача не зачтена.")
        notify_student(group_name, student_name, task_number, "Незачёт")
        move_submission_to_checked(group_name, student_name, task_number)
        start_message(call.message)
    else:
        bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="Немного подождите")
        if call.data == "Пособие_1":
            pdf_name = "С_разделяющимися_переменными.pdf"
        elif call.data == "Пособие_2":
            pdf_name = "Однородные_уравнения.pdf"
        elif call.data == "Пособие_3":
            pdf_name = "Линейные_уравнения_первого_порядка.pdf"
        elif call.data == "Пособие_4":
            pdf_name = "Уравнения_в_полных_дифференциалах.pdf"
        elif call.data == "Пособие_5":
            pdf_name = "Допускающие_понижение_порядка.pdf"
        elif call.data == "Пособие_6":
            pdf_name = "Однородные_относительно_искомой_функции.pdf"
        elif call.data == "Пособие_7":
            pdf_name = "Со_специальной_правой_частью.pdf"
        elif call.data == "Пособие_8":
            pdf_name = "Вариация_произвольных_постоянных.pdf"
        send_pdf(chat_id, manuals_folder, pdf_name)
        bot.delete_message(chat_id, message_id)
        start_message(call.message)


def process_task_number(message):
    chat_id = message.chat.id
    try:
        task_number = int(message.text.strip())
        if 1 <= task_number <= 15:
            user_state[chat_id]["task_number"] = task_number
            msg = bot.send_message(chat_id, "Отправьте файл с решением задачи:")
            bot.register_next_step_handler(msg, process_task_file)
        else:
            msg = bot.send_message(chat_id, "Неверный номер задачи. Пожалуйста, введите номер от 1 до 15:")
            bot.register_next_step_handler(msg, process_task_number)
    except ValueError:
        msg = bot.send_message(chat_id, "Неверный формат. Пожалуйста, введите номер от 1 до 15:")
        bot.register_next_step_handler(msg, process_task_number)

def process_task_file(message):
    chat_id = message.chat.id
    if message.content_type == 'document':
        task_number = user_state[chat_id]["task_number"]
        group = user_state[chat_id]["group"]
        name = user_state[chat_id]["name"]

        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        file_path = os.path.join("submissions", f"{group}_{name}_task_{task_number}.pdf")
        
        with open(file_path, 'wb') as new_file:
            new_file.write(downloaded_file)

        excel_file = os.path.join("groups", group + ".xlsx")
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == name:
                    cell = row[task_number + 1]
                    cell.value = "-"
                    pale_orange_fill = openpyxl.styles.PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
                    cell.fill = pale_orange_fill
                    wb.save(excel_file)
                    break

        bot.send_message(chat_id, "Файл успешно загружен и отмечен в таблице.")
        start_message(message)
    else:
        msg = bot.send_message(chat_id, "Неверный формат файла. Пожалуйста, отправьте файл с решением задачи:")
        bot.register_next_step_handler(msg, process_task_file)


def process_group_step(message):
    chat_id = message.chat.id
    group = message.text.strip()
    if group_exists(group):
        user_state[chat_id] = {"group": group}
        msg = bot.send_message(chat_id, "Введите ваше ФИО:")
        bot.register_next_step_handler(msg, process_name_step)
    else:
        msg = bot.send_message(chat_id, "Группа не найдена. Пожалуйста, введите правильную группу:")
        bot.register_next_step_handler(msg, process_group_step)

def process_name_step(message):
    chat_id = message.chat.id
    group = user_state[chat_id]["group"]
    name = message.text.strip()
    
    if name_exists_in_group(group, name):
        user_state[chat_id]["name"] = name
        save_user_state()
        
        excel_file = os.path.join("groups", group + ".xlsx")
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            assigned_variants = [row[1].value for row in sheet.iter_rows(min_row=2) if row[1].value is not None]
        else:
            assigned_variants = []
        
        if len(set(assigned_variants)) < 20:
            available_variants = [i for i in range(1, 21) if assigned_variants.count(i) < 1]
        else:
            available_variants = [i for i in range(1, 21) if assigned_variants.count(i) < 2]
        
        variant = random.choice(available_variants)
        update_variant(group, name, variant)
        
        pdf_name = f"Вариант_{variant}.pdf"
        send_pdf(chat_id, variants_folder, pdf_name)
        start_message(message)
    else:
        msg = bot.send_message(chat_id, "ФИО не найдено в группе. Пожалуйста, введите правильное ФИО:")
        bot.register_next_step_handler(msg, process_name_step)

def process_group_download(message):
    chat_id = message.chat.id
    group_name = message.text.strip()
    if group_name == "/exit_admin":
        handle_exit_admin(message)
    else:
        excel_file = os.path.join("groups", group_name + ".xlsx")
        if os.path.exists(excel_file):
            send_pdf(chat_id, "groups", group_name + ".xlsx")
        else:
            bot.send_message(chat_id, "Файл для данной группы не найден.")
        start_message(message)

def process_check_groups(message):
    groups = {}
    for filename in os.listdir(submissions_folder):
        match = re.match(r'([\w-]+)_([\w\s]+)_task_(\d+)\.pdf', filename, re.UNICODE)
        if match:
            group, student, task = match.groups()
            if group not in groups:
                groups[group] = set()
            groups[group].add(student)

    if groups:
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for group in groups.keys():
            keyboard.add(types.InlineKeyboardButton(text=group, callback_data=f"group_{group}"))
        bot.send_message(message.chat.id, "Выберите группу:", reply_markup=keyboard)
    else:
        bot.send_message(message.chat.id, "Нет новых задач для проверки.")
        start_message(message)

def process_check_students(message, group_name):
    chat_id = message.chat.id
    students = set()
    for filename in os.listdir(submissions_folder):
        match = re.match(rf'{group_name}_([\w\s]+)_task_(\d+)\.pdf', filename, re.UNICODE)
        if match:
            student = match.groups()[0]
            students.add(student)
    
    if students:
        current_selection[chat_id] = {'group': group_name}
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for student in students:
            keyboard.add(types.InlineKeyboardButton(text=student, callback_data=f"student_{student}"))
        bot.send_message(message.chat.id, "Выберите студента:", reply_markup=keyboard)
    else:
        bot.send_message(message.chat.id, "Нет новых задач для проверки у этой группы.")
        start_message(message)

def process_check_tasks(message, student_name):
    chat_id = message.chat.id
    group_name = current_selection[chat_id]['group']
    tasks = set()
    for filename in os.listdir(submissions_folder):
        match = re.match(rf'{group_name}_{student_name}_task_(\d+)\.pdf', filename, re.UNICODE)
        if match:
            task = match.groups()[0]
            tasks.add(task)
    
    if tasks:
        current_selection[chat_id]['student'] = student_name
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for task in tasks:
            keyboard.add(types.InlineKeyboardButton(text=f"Задача {task}", callback_data=f"task_{task}"))
        bot.send_message(message.chat.id, "Выберите задачу:", reply_markup=keyboard)
    else:
        bot.send_message(message.chat.id, "Нет новых задач для проверки у этого студента.")
        start_message(message)

def process_send_task(message, task_number):
    chat_id = message.chat.id
    group_name = current_selection[chat_id]['group']
    student_name = current_selection[chat_id]['student']
    file_name = f"{group_name}_{student_name}_task_{task_number}.pdf"
    file_path = os.path.join(submissions_folder, file_name)
    with open(file_path, 'rb') as f:
        bot.send_document(message.chat.id, f)
    
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    btn_pass = types.InlineKeyboardButton(text="Зачёт", callback_data=f"pass_{task_number}")
    btn_fail = types.InlineKeyboardButton(text="Незачёт", callback_data=f"fail_{task_number}")
    keyboard.add(btn_pass, btn_fail)
    bot.send_message(message.chat.id, "Поставить зачёт или незачёт?", reply_markup=keyboard)

def update_task_status(group, name, task_number, status):
    excel_file = os.path.join("groups", group + ".xlsx")
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == name:
                cell = row[int(task_number) + 1]
                cell.value = status
                pale_orange_fill = openpyxl.styles.PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
                cell.fill = pale_orange_fill
                wb.save(excel_file)
                break

def notify_student(group, student_name, task_number, verdict):
    for chat_id, data in user_state.items():
        if data.get("group") == group and data.get("name") == student_name:
            bot.send_message(chat_id, f"Ваша задача {task_number} получила оценку: {verdict}.")

def move_submission_to_checked(group, student_name, task_number):
    source_file = os.path.join(submissions_folder, f"{group}_{student_name}_task_{task_number}.pdf")
    dest_file = os.path.join(checked_folder, f"{group}_{student_name}_task_{task_number}.pdf")
    
    if os.path.exists(source_file):
        if os.path.exists(dest_file):
            base, ext = os.path.splitext(dest_file)
            counter = 1
            new_dest_file = f"{base}_{counter}{ext}"
            while os.path.exists(new_dest_file):
                counter += 1
                new_dest_file = f"{base}_{counter}{ext}"
            dest_file = new_dest_file
        
        os.rename(source_file, dest_file)


load_user_state()
load_admin_codes()
bot.polling(none_stop=True)
